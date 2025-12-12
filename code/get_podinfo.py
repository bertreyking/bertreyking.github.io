#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from kubernetes import client, config
import json
import re
from datetime import datetime

def get_pod_ips(pod):
    """从 Pod 的 annotations 中解析 multus IP 信息(k8s.v1.cni.cncf.io/network-status)"""
    ips = []

    # Pod.status.pod_ip 作为默认 IP（k8s 默认网卡）
    if pod.status.pod_ip:
        ips.append({"interface": "eth0", "ip": pod.status.pod_ip})

    ann = pod.metadata.annotations or {}
    net_status = ann.get("k8s.v1.cni.cncf.io/network-status")

    if net_status:
        try:
            net_status = json.loads(net_status)
            # 可能是 list，也可能是 dict
            if isinstance(net_status, dict):
                net_status = [net_status]

            for entry in net_status:
                if "ips" in entry:
                    for ip in entry["ips"]:
                        ips.append({
                            "interface": entry.get("interface", "unknown"),
                            "ip": ip
                        })
        except Exception as e:
            print(f"Failed to parse network-status for pod {pod.metadata.name}: {e}")

    return ips


def get_container_resources(container):
    """提取 container resources 信息"""
    res = container.resources
    requests = res.requests if res and res.requests else {}
    limits = res.limits if res and res.limits else {}

    return {
        "requests": {
            "cpu": requests.get("cpu"),
            "memory": requests.get("memory")
        },
        "limits": {
            "cpu": limits.get("cpu"),
            "memory": limits.get("memory")
        }
    }


def main():
    # 使用本地 kubeconfig
    config.load_kube_config(config_file=r"C:\code\python\k8s\config")

    v1 = client.CoreV1Api()
    pods = v1.list_pod_for_all_namespaces().items

    results = []

    for pod in pods:
        pod_info = {
            "namespace": pod.metadata.namespace,
            "name": pod.metadata.name,
            "node": pod.spec.node_name,
            "ips": get_pod_ips(pod),
            "containers": []
        }

        # 多容器处理
        for container in pod.spec.containers:
            pod_info["containers"].append({
                "name": container.name,
                "resources": get_container_resources(container)
            })

        results.append(pod_info)

    # 简单的 ASCII 表格渲染（无外部依赖）
    def format_table(headers, rows):
        # 计算每列宽度
        widths = []
        for i, h in enumerate(headers):
            col_width = len(str(h))
            for r in rows:
                col_width = max(col_width, len(str(r[i]) if i < len(r) else ""))
            widths.append(col_width)

        sep = "+" + "+".join(["-" * (w + 2) for w in widths]) + "+"
        header_line = "|" + "|".join([" " + str(headers[i]).ljust(widths[i]) + " " for i in range(len(headers))]) + "|"

        lines = [sep, header_line, sep]
        for r in rows:
            line = "|" + "|".join([" " + str(r[i]).ljust(widths[i]) + " " if i < len(r) else " " + "".ljust(widths[i]) + " " for i in range(len(headers))]) + "|"
            lines.append(line)
        lines.append(sep)
        return "\n".join(lines)

    def format_ips(ips):
        # 仅显示 eth0 和 net1，避免重复 IP
        order = ["eth0", "net1"]
        iface_map = {}
        seen_ips = set()

        for entry in ips:
            if not isinstance(entry, dict):
                continue
            iface = entry.get("interface") or "eth0"
            ip = entry.get("ip")
            if not ip:
                continue
            # 只关心 eth0 和 net1
            if iface not in order:
                continue
            # 去重相同 IP
            if ip in seen_ips:
                continue
            # 对每个 interface 仅保留第一个不同的 IP
            if iface not in iface_map:
                iface_map[iface] = ip
                seen_ips.add(ip)

        parts = []
        for iface in order:
            if iface in iface_map:
                parts.append(f"{iface}:{iface_map[iface]}")

        return ",".join(parts) if parts else "-"

    # 合并为一个表格：Namespace, Name, Node, IPs (interface:ip), Containers (names), Resources
    headers = ["Namespace", "Name", "Node", "IPs", "Containers", "Resources"]
    rows = []
    for p in results:
        ip_cell = format_ips(p.get("ips", []))
        container_names = ",".join([c.get("name") for c in p.get("containers", [])]) or "-"
        # 格式化容器 resources：支持多容器，request/limit 有则写入，否则写 none
        def format_resources(containers):
            parts = []
            for c in containers:
                name = c.get("name") or "-"
                res = c.get("resources") or {}
                req = res.get("requests") or {}
                lim = res.get("limits") or {}

                req_cpu = req.get("cpu") or "none"
                req_mem = req.get("memory") or "none"
                lim_cpu = lim.get("cpu") or "none"
                lim_mem = lim.get("memory") or "none"

                # 仅当内存格式为 <digits>m 时进行转换（例如 7301444403200m），其它格式保持原样
                def convert_m_to_mi(val):
                    if not isinstance(val, str):
                        return None
                    m = re.fullmatch(r"(\d+)m", val.strip())
                    if not m:
                        return None
                    try:
                        num = int(m.group(1))
                        # 解释为 milli-bytes -> bytes = num / 1000
                        bytes_val = num / 1000.0
                        mi = bytes_val / (1024 ** 2)
                        if mi >= 1:
                            return f"{int(mi)}Mi"
                        return f"{mi:.2f}Mi"
                    except Exception:
                        return None

                parsed_req_mem = convert_m_to_mi(req_mem) or (req_mem if req_mem != "none" else "none")
                parsed_lim_mem = convert_m_to_mi(lim_mem) or (lim_mem if lim_mem != "none" else "none")

                # 如果所有字段都是 none，则显示 name:none
                if req_cpu == req_mem == lim_cpu == lim_mem == "none":
                    parts.append(f"{name}:none")
                else:
                    parts.append(f"{name}:req(cpu={req_cpu},mem={parsed_req_mem});lim(cpu={lim_cpu},mem={parsed_lim_mem})")

            return "; ".join(parts) if parts else "-"

        res_cell = format_resources(p.get("containers", []))

        rows.append([
            p.get("namespace") or "-",
            p.get("name") or "-",
            p.get("node") or "-",
            ip_cell,
            container_names,
            res_cell,
        ])

    if rows:
        out = []
        out.append("Pods:")
        out.append(format_table(headers, rows))
        out.append("")
        output_text = "\n".join(out)

        # 写入按时间戳命名的 Excel 文件 podinfo_YYYYMMDD_HHMMSS.xlsx
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"podinfo_{ts}.xlsx"
        try:
            from openpyxl import Workbook

            wb = Workbook()
            ws = wb.active
            # 写入 header（首行）
            ws.append(headers)
            # 写入每一行数据
            for row in rows:
                ws.append(row)

            # 启用筛选（auto_filter）覆盖所有写入的列
            last_row = len(rows) + 1
            # 列范围，例如 A1:E{last_row}
            from openpyxl.utils import get_column_letter
            last_col = len(headers)
            last_col_letter = get_column_letter(last_col)
            ws.auto_filter.ref = f"A1:{last_col_letter}{last_row}"

            # 可选：自动调整列宽（简单实现）
            for col_idx, col in enumerate(ws.iter_cols(min_row=1, max_row=last_row, min_col=1, max_col=last_col), start=1):
                max_length = 0
                for cell in col:
                    try:
                        value = str(cell.value) if cell.value is not None else ""
                    except Exception:
                        value = ""
                    if len(value) > max_length:
                        max_length = len(value)
                col_letter = get_column_letter(col_idx)
                ws.column_dimensions[col_letter].width = min(50, max_length + 2)

            wb.save(filename)
            print(f"Written output to {filename}")
        except Exception as e:
            print(f"Failed to write Excel output to {filename}: {e}")


if __name__ == "__main__":
    main()

